import random
import string

import openpyxl

workbook=openpyxl.load_workbook("C://Users//91789//OneDrive//Desktop//testdadata.xlsx")
sheet=workbook["Sheet1"]
rows=sheet.max_row
columns=sheet.max_column
for r in range(1,rows+1):
    for j in range(1,columns+1):
        p=sheet.cell(row=r,column=j).value="".join(random.choice(string.ascii_lowercase) for i in range(10) )
        print(p,end=' ')
    print("")
    workbook.save("C://Users//91789//OneDrive//Desktop//testdadata.xlsx")


